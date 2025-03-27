import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font

# 定義供應商數據
suppliers = {
    "A": {"initial_price": 200000, "maintenance_rate": 0.22, "tax_rate": 0, "delivery_months": 3},
    "B": {"initial_price": 220000, "maintenance_rate": 0.20, "tax_rate": 0, "delivery_months": 2.5},
    "C": {"initial_price": 180000, "maintenance_rate": 0.25, "tax_rate": 0.05, "delivery_months": 4}
}

discount_rate = 0.05  # 折現率
years = 5
weights = {"價格": 0.4, "維護成本": 0.3, "交期": 0.3}  # 權重
data = {"供應商": [], "初始價格": [], "維護成本": [], "交期": [], "加權總分": []}

# 計算每項數據
for supplier, details in suppliers.items():
    initial_price = details["initial_price"]
    maintenance_cost = initial_price * details["maintenance_rate"]

    # 若含稅，計算去稅後的維護費用
    if details["tax_rate"] > 0:
        maintenance_cost /= (1 + details["tax_rate"])

    # 計算交期分數（以最短交期為基準）
    delivery_score = max(0, 4 - details["delivery_months"])  # 假設交期越短越好

    # 計算加權總分
    weighted_score = (
        (initial_price * weights["價格"]) +
        (maintenance_cost * (years - 1) * weights["維護成本"]) +
        (delivery_score * weights["交期"])
    )

    # 收集數據
    data["供應商"].append(supplier)
    data["初始價格"].append(initial_price)
    data["維護成本"].append(maintenance_cost * (years - 1))  # 累計4年維護費用
    data["交期"].append(delivery_score)
    data["加權總分"].append(weighted_score)

# 將數據寫入 Excel
df = pd.DataFrame(data)
excel_path = "Supplier_Cost_Breakdown.xlsx"

# 使用 openpyxl 寫入數據並格式化
wb = Workbook()
ws = wb.active
ws.title = "成本明細"

# 寫入數據
for r_idx, row in enumerate(df.itertuples(index=False), start=1):
    ws.append(row)

# 添加標題行
ws.insert_rows(1)
for col_idx, col_name in enumerate(df.columns, start=1):
    ws.cell(row=1, column=col_idx, value=col_name)

# 格式化數字為千分位並加上美元符號
for col in range(2, 5):  # 初始價格、維護成本、交期
    for row in range(2, len(data["供應商"]) + 2):  # 跳過標題行
        cell = ws.cell(row=row, column=col)
        cell.number_format = '"$"#,##0'

# 加粗標題行
for col in range(1, 6):
    cell = ws.cell(row=1, column=col)
    cell.font = Font(bold=True)

# 創建條形圖
chart = BarChart()
chart.title = "供應商成本明細"
chart.x_axis.title = "供應商"
chart.y_axis.title = "成本與分數"

data_ref = Reference(ws, min_col=2, min_row=1, max_col=5, max_row=len(data["供應商"]) + 1)
categories_ref = Reference(ws, min_col=1, min_row=2, max_row=len(data["供應商"]) + 1)
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(categories_ref)

# 將圖表添加到工作表
ws.add_chart(chart, "G2")

# 儲存 Excel
wb.save(excel_path)

print(f"數據已成功導出至 {excel_path}，並生成圖表。")